"""
Export simulation results to Excel file
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
import os
import json


def extract_text_from_json_response(json_str):
    """
    Extract and combine all text from JSON API response.
    
    Args:
        json_str: JSON string of API response
        
    Returns:
        str: Combined text from all text items, separated by spaces
    """
    if not json_str or json_str.strip() == "":
        return ""
    
    try:
        api_response = json.loads(json_str)
        text_list = api_response.get("text", [])
        if not isinstance(text_list, list) or len(text_list) == 0:
            return ""
        
        # Extract text from each item
        extracted_texts = []
        for item in text_list:
            if isinstance(item, dict):
                # Object with 'text' property
                text = item.get("text", "")
                if text:
                    extracted_texts.append(text)
            elif isinstance(item, str):
                # String
                extracted_texts.append(item)
        
        # Join all texts with space
        return " ".join(extracted_texts).strip()
    except (json.JSONDecodeError, AttributeError) as e:
        return ""

def export_to_excel(
    message_history: List[Dict[str, str]],
    response_times: List[float],
    full_logs: List[str],
    output_dir: str = "results",
    filename: str = None,
    api_base_url: str = "http://103.253.20.30:9404",
    bot_id: Optional[int] = None,
    conversation_id: Optional[str] = None
) -> str:
    """
    Export simulation results to Excel file with columns: Role, curl API, raw_output, text_output, Response time.
    
    Args:
        message_history: List of messages with 'role' and 'content' keys
        response_times: List of response times in seconds
        full_logs: List of full API response logs (JSON strings)
        output_dir: Directory to save the Excel file
        filename: Optional filename. If None, will generate with timestamp
        api_base_url: Base URL of the API (for generating cURL commands)
        bot_id: Bot ID (for generating cURL commands)
        conversation_id: Conversation ID (for generating cURL commands)
        
    Returns:
        str: Path to the created Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulation_results_{timestamp}.xlsx"
    
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    filepath = os.path.join(output_dir, filename)
    
    # Prepare data for Excel
    excel_data = []
    
    for i, msg in enumerate(message_history):
        role = msg["role"]
        content = msg["content"]
        response_time = response_times[i] if i < len(response_times) else 0
        
        # Determine Role name
        role_name = "RoleA" if role == "roleA" else "RoleB"
        
        # Generate cURL command
        curl_command = ""
        if role == "roleA":
            # RoleA uses OpenAI API
            if i == 0:
                # Initial message - no API call
                curl_command = "Initial message"
            else:
                # OpenAI API call
                curl_command = "OpenAI API (Chat Completion)"
        elif role == "roleB":
            # RoleB uses HTTP API - generate cURL command
            # The cURL should show the request that was sent to get this response
            if conversation_id and api_base_url:
                webhook_endpoint = f"{api_base_url}/robot-ai-lesson/api/v1/bot/webhook"
                
                # Find the message that was sent to get this RoleB response
                # RoleB responses are based on the previous RoleA message
                message_to_send = ""
                
                # Look backwards to find the most recent RoleA message
                for j in range(i - 1, -1, -1):
                    if message_history[j]["role"] == "roleA":
                        message_to_send = message_history[j]["content"]
                        break
                
                # If no RoleA message found (should not happen), use default
                if not message_to_send:
                    message_to_send = "sẵn sàng"
                
                payload = {
                    "conversation_id": conversation_id,
                    "message": message_to_send
                }
                
                # Create cURL command - escape properly for Excel
                payload_json = json.dumps(payload, ensure_ascii=False)
                # For Excel, use single line format that's easier to read
                curl_command = f'curl -X POST "{webhook_endpoint}" -H "Content-Type: application/json" -d \'{payload_json}\''
            else:
                curl_command = "API Call (details not available)"
        
        # Get raw_output and text_output
        raw_output = ""
        text_output = content  # Default to content from message_history
        
        if role == "roleB":
            # For RoleB, get raw JSON and extract text
            if i < len(full_logs) and full_logs[i]:
                raw_output = full_logs[i]
                # Extract text from JSON
                text_output = extract_text_from_json_response(raw_output)
                # If extraction failed, use content from message_history
                if not text_output:
                    text_output = content
        else:
            # For RoleA, raw_output is empty (OpenAI API)
            raw_output = ""
            text_output = content
        
        # Add row to Excel data
        excel_data.append({
            "Role": role_name,
            "curl API": curl_command,
            "raw_output": raw_output,
            "text_output": text_output,
            "Response time": f"{response_time:.6f}"
        })
    
    # Create DataFrame
    df = pd.DataFrame(excel_data)
    
    # Write to Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Simulation Results')
        
        # Get the worksheet to adjust column widths
        worksheet = writer.sheets['Simulation Results']
        
        # Set column widths for better readability
        worksheet.column_dimensions['A'].width = 10  # Role
        worksheet.column_dimensions['B'].width = 80  # curl API
        worksheet.column_dimensions['C'].width = 100  # raw_output (JSON)
        worksheet.column_dimensions['D'].width = 80  # text_output
        worksheet.column_dimensions['E'].width = 15  # Response time
        
        # Enable text wrapping for raw_output column (column C)
        from openpyxl.styles import Alignment
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"\n✅ Excel file created successfully: {filepath}")
    return filepath


def export_multiple_simulations_to_excel(
    simulation_results: List[Dict[str, Any]],
    output_dir: str = "results",
    filename: str = None
) -> str:
    """
    Export multiple simulation results to a single Excel file with ONE sheet.
    All simulations will be combined into a single sheet with a "Simulation" column.
    
    Args:
        simulation_results: List of simulation result dictionaries, each containing:
            - message_history: List of messages
            - response_times: List of response times
            - full_logs: List of full API response logs
            - api_client: API client object (for base_url, bot_id, conversation_id)
            - simulation_name: Name of the simulation (optional)
        output_dir: Directory to save the Excel file
        filename: Optional filename. If None, will generate with timestamp
        
    Returns:
        str: Path to the created Excel file
    """
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"simulations_{timestamp}.xlsx"
    
    # Ensure filename has .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    filepath = os.path.join(output_dir, filename)
    
    # Import openpyxl styles
    from openpyxl.styles import Alignment
    
    # Collect all data from all simulations into one list
    all_excel_data = []
    
    for idx, result in enumerate(simulation_results):
        message_history = result["message_history"]
        response_times = result["response_times"]
        full_logs = result["full_logs"]
        api_client = result["api_client"]
        
        # Get simulation name
        simulation_name = result.get("simulation_name", f"Simulation_{idx+1}")
        
        # Prepare data for Excel
        for i, msg in enumerate(message_history):
            role = msg["role"]
            content = msg["content"]
            response_time = response_times[i] if i < len(response_times) else 0
            
            # Determine Role name
            role_name = "RoleA" if role == "roleA" else "RoleB"
            
            # Generate cURL command
            curl_command = ""
            if role == "roleA":
                # RoleA uses OpenAI API
                if i == 0:
                    # Initial message - no API call
                    curl_command = "Initial message"
                else:
                    # OpenAI API call
                    curl_command = "OpenAI API (Chat Completion)"
            elif role == "roleB":
                # RoleB uses HTTP API - generate cURL command
                conversation_id = api_client.current_conversation_id
                api_base_url = api_client.base_url
                
                if conversation_id and api_base_url:
                    webhook_endpoint = f"{api_base_url}/robot-ai-lesson/api/v1/bot/webhook"
                    
                    # Find the message that was sent to get this RoleB response
                    message_to_send = ""
                    
                    # Look backwards to find the most recent RoleA message
                    for j in range(i - 1, -1, -1):
                        if message_history[j]["role"] == "roleA":
                            message_to_send = message_history[j]["content"]
                            break
                    
                    # If no RoleA message found (should not happen), use default
                    if not message_to_send:
                        message_to_send = "sẵn sàng"
                    
                    payload = {
                        "conversation_id": conversation_id,
                        "message": message_to_send
                    }
                    
                    # Create cURL command
                    payload_json = json.dumps(payload, ensure_ascii=False)
                    curl_command = f'curl -X POST "{webhook_endpoint}" -H "Content-Type: application/json" -d \'{payload_json}\''
                else:
                    curl_command = "API Call (details not available)"
            
            # Get raw_output and text_output
            raw_output = ""
            text_output = content  # Default to content from message_history
            
            if role == "roleB":
                # For RoleB, get raw JSON and extract text
                if i < len(full_logs) and full_logs[i]:
                    raw_output = full_logs[i]
                    # Extract text from JSON
                    text_output = extract_text_from_json_response(raw_output)
                    # If extraction failed, use content from message_history
                    if not text_output:
                        text_output = content
            else:
                # For RoleA, raw_output is empty (OpenAI API)
                raw_output = ""
                text_output = content
            
            # Add row to Excel data with Simulation column
            all_excel_data.append({
                "Simulation": simulation_name,
                "Role": role_name,
                "curl API": curl_command,
                "raw_output": raw_output,
                "text_output": text_output,
                "Response time": f"{response_time:.6f}"
            })
    
    # Create DataFrame with all simulations combined
    df = pd.DataFrame(all_excel_data)
    
    # Write to Excel - ONE sheet only
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='All Simulations')
        
        # Get the worksheet to adjust column widths
        worksheet = writer.sheets['All Simulations']
        
        # Set column widths for better readability
        worksheet.column_dimensions['A'].width = 20  # Simulation
        worksheet.column_dimensions['B'].width = 10  # Role
        worksheet.column_dimensions['C'].width = 80  # curl API
        worksheet.column_dimensions['D'].width = 100  # raw_output (JSON)
        worksheet.column_dimensions['E'].width = 80  # text_output
        worksheet.column_dimensions['F'].width = 15  # Response time
        
        # Enable text wrapping for raw_output column (column D)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"\n✅ Excel file created successfully with 1 sheet containing {len(simulation_results)} simulations: {filepath}")
    return filepath